"""Excel 输入解析 + 结果导出（不匹配行高亮）。

输入：
- 发票：列 `pic_url`(必，图片直链)，`supply_money`(选，原始金额标签)。对齐 data/发票数据.xlsx。
- 停车：列 `cost_images`(必，逗号分隔的图片直链)，`cost`/`vin`(选，标签)。对齐 data/停车数据.xlsx。

导出：原始/识别字段一张表，`amount_match`/`vin_match` 为 ✗ 的行整行红色高亮。
"""
from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill

_MISS_FILL = PatternFill(start_color="FFF8CBAD", end_color="FFF8CBAD", fill_type="solid")  # 浅红
_HEAD_FILL = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")  # 深蓝表头
_HEAD_FONT = Font(color="FFFFFFFF", bold=True)


def read_invoice_excel(data: bytes) -> list[dict]:
    df = pd.read_excel(BytesIO(data))
    cols = {c.lower().strip(): c for c in df.columns}
    if "pic_url" not in cols:
        raise ValueError("发票 Excel 需包含列 pic_url（图片直链）")
    url_c = cols["pic_url"]
    amt_c = cols.get("supply_money")
    rows = []
    for _, r in df.iterrows():
        u = r[url_c]
        if pd.isna(u) or not str(u).strip():
            continue
        label = None if (amt_c is None or pd.isna(r[amt_c])) else float(r[amt_c])
        rows.append({"pic_url": str(u).strip(), "supply_money": label})
    return rows


def read_parking_excel(data: bytes) -> list[dict]:
    df = pd.read_excel(BytesIO(data))
    cols = {c.lower().strip(): c for c in df.columns}
    if "cost_images" not in cols:
        raise ValueError("停车 Excel 需包含列 cost_images（逗号分隔的图片直链）")
    img_c = cols["cost_images"]
    cost_c, vin_c, wf_c = cols.get("cost"), cols.get("vin"), cols.get("workflow_no")
    rows = []
    for _, r in df.iterrows():
        raw = r[img_c]
        if pd.isna(raw) or not str(raw).strip():
            continue
        urls = [u.strip() for u in str(raw).split(",") if u.strip()]
        rows.append({
            "workflow_no": None if (wf_c is None or pd.isna(r[wf_c])) else str(r[wf_c]),
            "images": urls,
            "cost": None if (cost_c is None or pd.isna(r[cost_c])) else float(r[cost_c]),
            "vin": None if (vin_c is None or pd.isna(r[vin_c])) else str(r[vin_c]),
        })
    return rows


def results_to_xlsx(results: list[dict], columns: list[str] | None = None,
                    flag_cols: tuple[str, ...] = ("amount_match", "vin_match")) -> bytes:
    """结果列表 → 高亮 xlsx 字节。flag_cols 中任一为 '✗' 的行整行标红。"""
    df = pd.DataFrame(results)
    if columns:
        # 仅保留约定列顺序（丢弃 raw 等冗余字段，导出干净）
        df = df.reindex(columns=[c for c in columns if c in df.columns])
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="识别结果")
        ws = writer.sheets["识别结果"]
        headers = [c.value for c in ws[1]]
        flag_idx = [i for i, h in enumerate(headers) if h in flag_cols]
        for cell in ws[1]:
            cell.fill, cell.font = _HEAD_FILL, _HEAD_FONT
        for row in ws.iter_rows(min_row=2):
            if any(row[i].value == "✗" for i in flag_idx):
                for cell in row:
                    cell.fill = _MISS_FILL
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 60)
    return bio.getvalue()
